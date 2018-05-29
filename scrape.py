# Hockey data scraping test code
# PC updates using PB's code as reference
# Reference for beautifulsoup & urllib:
# [1] https://medium.freecodecamp.org/how-to-scrape-websites-with-python-and-beautifulsoup-5946935d93fe

# import block
from bs4 import BeautifulSoup
import urllib.request
from urllib.request import urlopen
import numpy as np
import matplotlib as plt
import xlrd, xlwt
import timeit as time
import openpyxl as xcel

#book = xlrd.open_workbook('games.xlsx')
#game_sheet = book.sheet_by_index(0)
#colGames = game_sheet.col_slice(colx=0,start_rowx=1,end_rowx=82)
#numGames = len(colGames)
wb = xcel.load_workbook('games.xlsx')
ws = wb.active
start = time.default_timer()

hrUrl = 'https://www.hockey-reference.com/teams/TOR/'
yearCtr = 2016
position = 0
#gamesDict = {}
#games[2018] = []
#games[2018].append(2)
i = 0
while yearCtr < 2019:
    allGames = hrUrl + str(yearCtr) + '_games.html'
    print(allGames)
    gamesPage = urlopen(allGames)
    excelPop = BeautifulSoup(gamesPage,'html.parser')


    #This checks to see if the excel sheet has a sheet named gamedays and creates it if it doesnt
    gameDaysName = 'gamedays'
    if gameDaysName in wb.sheetnames:
        ws1 = wb[gameDaysName]
    else:
        ws1 = wb.create_sheet(gameDaysName)

    #amountOfGames = excelPop.find_all('th',{'data-stat': "games"})

    #this iterates over the bs4'd hockey reference page and grabs all href links that match the criteria
    for a in excelPop.find_all('a', href=True):
        if a['href'][0:11] == '/boxscores/':
              if len(a['href'][11:-5]) > 11:
                ws1.cell(row=i + 1 + position, column=1).value = a['href'][11:-5]
                i += 1
    # i counts how many games were parsed
    numOfGames = i - 1
    yearCtr += 1
    position += 37

ctr = 0
if yearCtr == 2018:
    numOfGames = numOfGames - 1
    while ctr < numOfGames:
        ws1.cell(row=ctr + 1 + position, column=1).value = ws1.cell(row=ctr + 2, column=1).value
        ctr += 1
    ws1.cell(row=ctr + 1, column=1).value = ""

wb.save("games.xlsx")

position = 0
print("there are " + str(numOfGames) + " games")
for gameNumber in range(1, 500):
    cell_id = 'A' + str(gameNumber)
    gameHeader = ws[cell_id].value
    if gameHeader is None:
        gameNumber = 0
        continue
    print("game num " + str(gameNumber) + " is at position " + str(position))
    position += 1
    #print("Currently parsing game " + str(gameNumber) + " position of " + str(position) + " cell of " + cell_id + " gameheader of " + gameHeader)
    sitebase = 'https://www.hockey-reference.com/boxscores/'
    htmlTag = '.html'
    siteOne = sitebase + gameHeader + htmlTag
    print(siteOne)
    page = urlopen(siteOne)
    # Parse the html using beautiful soup and store in variable "soup" [1]
    # is there a way to automatically iterate within the html format?
    # 17 relevant columns in this table

    soup = BeautifulSoup(page,'html.parser')
    container = soup.find('tbody') # body that contains table data (numbers)
    players = soup.find_all('td',{'data-stat': "player"})
    goals = soup.find_all('td',{'data-stat': "goals"})
    assists = soup.find_all('td',{'data-stat': "assists"})
    points = soup.find_all('td',{'data-stat': "points"})
    plus_minus = soup.find_all('td',{'data-stat': "plus_minus"})
    pen_min = soup.find_all('td',{'data-stat': "pen_min"})
    goals_ev = soup.find_all('td',{'data-stat': "goals_ev"})
    goals_pp = soup.find_all('td',{'data-stat': "goals_pp"})
    goals_sh = soup.find_all('td',{'data-stat': "goals_sh"})
    goals_gw = soup.find_all('td',{'data-stat': "goals_gw"})
    assists_ev = soup.find_all('td',{'data-stat': "assists_ev"})
    assists_sh = soup.find_all('td',{'data-stat': "assists_sh"})
    assists_pp = soup.find_all('td',{'data-stat': "assists_pp"})
    shots = soup.find_all('td',{'data-stat': "shots"})
    shot_pct = soup.find_all('td',{'data-stat': "shot_pct"})
    shifts = soup.find_all('td',{'data-stat': "shifts"})
    time_on_ice = soup.find_all('td',{'data-stat': "time_on_ice"})
    # Will add advanced stats, later..
    numCol = 17 # number of stats collected above
    numRow = len(players) # number of players in the game - includes goalies and opposing team players

    #print(numRow)
    # ^ automate above in forloop? Can that be done with html mumbo jumbo?
    # Why do you want to print this ^^ stuff?

    # checks if the sheets exist that align with the date. If it exists it updates it, otherwise it creates a new sheet
    if gameHeader in wb.sheetnames:
        ws1 = wb[gameHeader]
    else:
        ws1 = wb.create_sheet(gameHeader)

    # Added column headers
    # I limited the range to 20 to grab all the toronto skaters and avoid the non equal dataset error
    for i in range(1,20):
        ws1.cell(row=1, column=1).value = "Name"
        ws1.cell(row=1, column=2).value = "Goals"
        ws1.cell(row=1, column=3).value = "Assists"
        ws1.cell(row=1, column=4).value = "Points"
        ws1.cell(row=1, column=5).value = "+/-"
        ws1.cell(row=1, column=6).value = "PIMS"
        ws1.cell(row=1, column=7).value = "Goals EV"
        ws1.cell(row=1, column=8).value = "Goals PP"
        ws1.cell(row=1, column=9).value = "Goals SH"
        ws1.cell(row=1, column=10).value = "Goals GW"
        ws1.cell(row=1, column=11).value = "Assists EV"
        ws1.cell(row=1, column=12).value = "Assists SH"
        ws1.cell(row=1, column=13).value = "Assists PP"
        ws1.cell(row=1, column=14).value = "Shots"
        ws1.cell(row=1, column=15).value = "Shot %"
        ws1.cell(row=1, column=16).value = "Shifts"
        ws1.cell(row=1, column=17).value = "TOI"

        if siteOne[-8:-5] == "TOR":
            #This print statement is just here so I could find the correct arr positions
            #print("Player: " + players[i-21].text + " has " + goals[i-20].text + " goals " + assists[i-20].text + " assists " + points[i-20].text + " points")

            ws1.cell(row=i+1, column=1).value = players[i - 21].text
            ws1.cell(row=i+1, column=2).value = goals[i-20].text
            ws1.cell(row=i+1, column=3).value = assists[i-20].text
            ws1.cell(row=i+1, column=4).value = points[i-20].text
            ws1.cell(row=i+1, column=5).value = plus_minus[i-20].text
            ws1.cell(row=i+1, column=6).value = pen_min[i-21].text
            ws1.cell(row=i+1, column=7).value = goals_ev[i-20].text
            ws1.cell(row=i+1, column=8).value = goals_pp[i-20].text
            ws1.cell(row=i+1, column=9).value = goals_sh[i-20].text
            ws1.cell(row=i+1, column=10).value = goals_gw[i-20].text
            ws1.cell(row=i+1, column=11).value = assists_ev[i-20].text
            ws1.cell(row=i+1, column=12).value = assists_sh[i-20].text
            ws1.cell(row=i+1, column=13).value = assists_pp[i-20].text
            ws1.cell(row=i+1, column=14).value = shots[i-20].text
            ws1.cell(row=i+1, column=15).value = shot_pct[i-20].text
            ws1.cell(row=i+1, column=16).value = shifts[i-20].text
            ws1.cell(row=i+1, column=17).value = time_on_ice[i-21].text
        else:
            #print("Player: " + players[i - 1].text + " has " + goals[i - 1].text + " goals " + assists[i - 1].text + " assists " + points[i - 1].text + " points")

            ws1.cell(row=i+1, column=1).value = players[i - 1].text
            ws1.cell(row=i+1, column=2).value = goals[i - 1].text
            ws1.cell(row=i+1, column=3).value = assists[i - 1].text
            ws1.cell(row=i+1, column=4).value = points[i - 1].text
            ws1.cell(row=i+1, column=5).value = plus_minus[i - 1].text
            ws1.cell(row=i+1, column=6).value = pen_min[i - 1].text
            ws1.cell(row=i+1, column=7).value = goals_ev[i - 1].text
            ws1.cell(row=i+1, column=8).value = goals_pp[i - 1].text
            ws1.cell(row=i+1, column=9).value = goals_sh[i - 1].text
            ws1.cell(row=i+1, column=10).value = goals_gw[i - 1].text
            ws1.cell(row=i+1, column=11).value = assists_ev[i - 1].text
            ws1.cell(row=i+1, column=12).value = assists_sh[i - 1].text
            ws1.cell(row=i+1, column=13).value = assists_pp[i - 1].text
            ws1.cell(row=i+1, column=14).value = shots[i - 1].text
            ws1.cell(row=i+1, column=15).value = shot_pct[i - 1].text
            ws1.cell(row=i+1, column=16).value = shifts[i - 1].text
            ws1.cell(row=i+1, column=17).value = time_on_ice[i-1].text

wb.save("games.xlsx")
stop = time.default_timer()
print("Runtime in seconds: ", stop-start)

# Note: When the HTML-Leafs-only issue is resolved, there will be no definite way to know who they played for that game
# ... Maybe look for an "opponent" tag (or equivalent" and save it somewhere in the header?
# Have Row 1 = Leafs VS Opponent, Date, Attendance, other misc info
# Row 2 = Column Headers
# Row 3 and on: Leafs stats corresponding to ^